import streamlit as st
import requests
import pandas as pd
import io
import os
import glob
import logging
from datetime import datetime, date
from enum import Enum
from typing import Dict, Any, Union, Optional

# Imports para Excel
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Configuração de Logs
logger = logging.getLogger(__name__)

# Dependências Opcionais
try: 
    from fpdf import FPDF
except ImportError: 
    logger.warning("Biblioteca 'fpdf' não encontrada. Funcionalidade de PDF desativada.")
    FPDF = None

try: 
    import pypdf
except ImportError: 
    pypdf = None

class TransactionType(Enum):
    INCOME = "Receita"
    EXPENSE = "Despesa"
    INVESTMENT = "Investimento"

class KnowledgeBaseLoader:
    @staticmethod
    def _read_pdf(file_path: str) -> str:
        if not pypdf: return "[ERRO] Biblioteca 'pypdf' não instalada."
        text = ""
        try:
            reader = pypdf.PdfReader(file_path)
            for page in reader.pages:
                extracted = page.extract_text()
                if extracted: text += extracted + "\n"
            return f"\n--- CONTEÚDO PDF ({os.path.basename(file_path)}) ---\n{text}\n"
        except Exception as e:
            return f"[ERRO LEITURA PDF]: {e}\n"

    @staticmethod
    @st.cache_data(ttl=3600)
    def load_knowledge(source: str) -> str:
        if not source: return ""
        combined_text = ""
        try:
            if source.startswith("http"):
                try:
                    r = requests.get(source, timeout=10)
                    if r.status_code == 200: return f"--- WEB ({source}) ---\n{r.text}\n"
                    return f"[ERRO URL]: {r.status_code}"
                except Exception as e: return f"[ERRO CONEXÃO]: {e}"
            elif os.path.isdir(source):
                files = glob.glob(os.path.join(source, '*'))
                valid = [f for f in files if f.lower().endswith(('.txt', '.md', '.pdf'))]
                for f_path in valid:
                    try:
                        if f_path.lower().endswith('.pdf'): combined_text += KnowledgeBaseLoader._read_pdf(f_path)
                        else:
                            with open(f_path, "r", encoding="utf-8") as f: combined_text += f"\n--- ARQ ({os.path.basename(f_path)}) ---\n{f.read()}\n"
                    except: continue
                return combined_text
            elif os.path.exists(source):
                if source.lower().endswith('.pdf'): return KnowledgeBaseLoader._read_pdf(source)
                with open(source, "r", encoding="utf-8") as f: return f"--- REGRAS ---\n{f.read()}\n"
            return f"[AVISO] Fonte não encontrada: {source}"
        except Exception as e: return f"[ERRO CRÍTICO]: {e}"

class DomainValidators:
    @staticmethod
    def validate_amount(amount: Any) -> float:
        try:
            val = float(amount)
            if val <= 0: raise ValueError("Valor deve ser maior que zero.")
            return val
        except: raise ValueError(f"Valor inválido: {amount}")

    @staticmethod
    def normalize_type(type_str: str) -> str:
        if not type_str: return TransactionType.EXPENSE.value
        t = str(type_str).strip().lower()
        if t in ['expense', 'outcome', 'gasto', 'saída', 'despesa', 'debit']: return TransactionType.EXPENSE.value
        elif t in ['income', 'entry', 'ganho', 'entrada', 'receita', 'credit']: return TransactionType.INCOME.value
        return TransactionType.EXPENSE.value

    @staticmethod
    def validate_date(date_val: Union[str, date, datetime]) -> str:
        if not date_val: return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(date_val, datetime): return date_val.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(date_val, date): return date_val.strftime('%Y-%m-%d')
        return str(date_val)

@st.cache_data(ttl=300, show_spinner=False)
def get_market_data() -> Dict[str, Any]:
    data = {"USD": 5.40, "EUR": 6.30, "GBP": 7.25, "BTC": 520000.0, "status": "offline"}
    headers = {"User-Agent": "SmartWallet/1.0"}
    try:
        if "FXRATES_KEY" in st.secrets:
            url = f"https://api.fxratesapi.com/latest?base=USD&currencies=BRL,EUR,GBP,BTC&api_key={st.secrets['FXRATES_KEY']}"
            r = requests.get(url, headers=headers, timeout=5)
            if r.status_code == 200:
                d = r.json()
                if d.get('success'):
                    rt = d.get('rates', {})
                    usd = float(rt.get('BRL', 5.40))
                    data.update({"USD": usd, "status": "online (FX)"})
                    if rt.get('EUR'): data['EUR'] = usd / float(rt['EUR'])
                    if rt.get('GBP'): data['GBP'] = usd / float(rt['GBP'])
                    if rt.get('BTC'): data['BTC'] = usd / float(rt['BTC'])
                    return data
    except: pass
    try:
        r = requests.get("https://economia.awesomeapi.com.br/last/USD-BRL,EUR-BRL,GBP-BRL,BTC-BRL", headers=headers, timeout=5)
        if r.status_code == 200:
            d = r.json()
            data["USD"] = float(d.get('USDBRL', {}).get('bid', data["USD"]))
            data["EUR"] = float(d.get('EURBRL', {}).get('bid', data["EUR"]))
            data["GBP"] = float(d.get('GBPBRL', {}).get('bid', data["GBP"]))
            data["BTC"] = float(d.get('BTCBRL', {}).get('bid', data["BTC"]))
            data["status"] = "online (API)"
            return data
    except: pass
    return data

class DocGenerator:
    @staticmethod
    def to_excel(df: pd.DataFrame) -> io.BytesIO:
        out = io.BytesIO()
        try:
            if df.empty: return out
            d = df.drop(columns=['id', 'proof_data', 'proof_name'], errors='ignore').copy()
            d['date'] = pd.to_datetime(d['date'], errors='coerce').dt.strftime('%d/%m/%Y %H:%M')
            d['amount'] = pd.to_numeric(d['amount'], errors='coerce').fillna(0.0)
            d = d.rename(columns={'date':'Data','amount':'Valor','category':'Categoria','description':'Desc','type':'Tipo'})
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                d.to_excel(writer, index=False, sheet_name="Extrato")
                ws = writer.sheets["Extrato"]
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                for col in ws.columns:
                    ws.column_dimensions[get_column_letter(col[0].column)].width = 20
            out.seek(0)
        except: return io.BytesIO()
        return out

    @staticmethod
    def to_pdf(user: str, df: pd.DataFrame, inc: float, exp: float, bal: float, period: str) -> Optional[bytes]:
        if FPDF is None: return None
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Título
            pdf.set_font("Arial", 'B', 16)
            pdf.set_text_color(76, 175, 80)
            pdf.cell(0, 10, "SmartWallet Relatorio", ln=True, align='C')
            
            # Subtítulo (Sanitizado)
            pdf.set_font("Arial", '', 10)
            pdf.set_text_color(50)
            safe_period = str(period).encode('latin-1', 'ignore').decode('latin-1')
            pdf.cell(0, 10, f"Periodo: {safe_period}", ln=True, align='C')
            
            # Resumo
            pdf.set_fill_color(240)
            pdf.rect(10, 30, 190, 15, 'F')
            pdf.set_y(32)
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(63, 10, f"Entradas: {inc:,.2f}", align='C')
            pdf.cell(63, 10, f"Saidas: {exp:,.2f}", align='C')
            pdf.cell(63, 10, f"Saldo: {bal:,.2f}", align='C')
            
            pdf.ln(20)
            
            # Cabeçalho Tabela
            cols = [("Data", 30), ("Tipo", 25), ("Categoria", 40), ("Desc", 50), ("Valor", 45)]
            pdf.set_fill_color(50); pdf.set_text_color(255)
            for c, w in cols: pdf.cell(w, 8, c, 1, 0, 'C', True)
            pdf.ln()
            
            # Linhas
            pdf.set_text_color(0); pdf.set_font("Arial", '', 9)
            for _, r in df.head(100).iterrows():
                s_date = str(r['date'])[:10]
                s_type = str(r['type']).encode('latin-1', 'ignore').decode('latin-1')
                s_cat = str(r['category'])[:18].encode('latin-1', 'ignore').decode('latin-1')
                s_desc = str(r['description'])[:25].encode('latin-1', 'ignore').decode('latin-1')
                s_val = f"{float(r['amount']):,.2f}"
                
                pdf.cell(30, 8, s_date, 1)
                pdf.cell(25, 8, s_type, 1)
                pdf.cell(40, 8, s_cat, 1)
                pdf.cell(50, 8, s_desc, 1)
                pdf.cell(45, 8, s_val, 1, 0, 'R')
                pdf.ln()
            
            return pdf.output(dest='S').encode('latin-1', 'ignore')
        except Exception as e:
            print(f"ERRO PDF: {e}")
            return None
